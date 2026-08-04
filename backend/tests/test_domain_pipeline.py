from __future__ import annotations

import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

from app.domain.models import (
    CompanySettings,
    MappingProfile,
    MappingStatus,
    MaterialityProfile,
    VarianceProfile,
    VarianceThreshold,
)
from app.domain.repository import InMemoryRepository
from app.services.event_engine import construct_event
from app.services.import_pipeline import (
    SettlementRow,
    normalize_general_ledger,
    normalize_settlement,
)
from app.services.mapping import propose_mapping
from app.services.orchestrator import process_journals
from app.services.variance import analyze_variance, monthly_values


class MappingAndPipelineTest(unittest.TestCase):
    def make_settlement(self) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["계정코드", "계정명", "당기"])
        sheet.append([122500, "현금", 501000000])
        sheet.append([344000, "자본금", -500000])
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        return Path(handle.name)

    def test_settlement_three_column_form_uses_current_period_as_amount(self) -> None:
        path = self.make_settlement()
        try:
            proposal = propose_mapping(path, "SETTLEMENT_SCHEDULE")
            self.assertEqual(proposal.mapping["account_code"], "계정코드")
            self.assertEqual(proposal.mapping["account_name"], "계정명")
            self.assertEqual(proposal.mapping["amount"], "당기")
            self.assertEqual(proposal.mapping["period"], "__UPLOAD_PERIOD__")
            self.assertFalse(proposal.missing_required)

            company = CompanySettings("P001", "테스트회사", "소재")
            profile = MappingProfile(
                company_id=company.id,
                source_type="SETTLEMENT_SCHEDULE",
                sheet_name=proposal.sheet_name,
                header_row=proposal.header_row,
                source_signature=proposal.signature,
                mapping=proposal.mapping,
                status=MappingStatus.APPROVED,
            )
            rows = normalize_settlement(path, profile, upload_period="2026-07")
            self.assertEqual(rows[0].period, "2026-07")
            self.assertEqual(rows[0].amount, Decimal("501000000"))
        finally:
            path.unlink(missing_ok=True)

    def make_gl(self) -> Path:
        workbook = Workbook()
        workbook.active.title = "Sheet1"
        workbook.create_sheet("Sheet2")
        sheet = workbook.create_sheet("Sheet3")
        sheet.append(
            [
                "회사코드",
                "전표번호",
                "전기일자",
                "회계연도",
                "회계월",
                "전표적요",
                "전표적요상세",
                "차대변구분자",
                "전표금액(기준통화)",
                "계정과목코드",
                "계정과목명",
            ]
        )
        sheet.append(["P001", "JE-1", date(2026, 7, 31), 2026, 7, "개발비", "양극재 개발", "차변", 1000000000, "120100", "개발비"])
        sheet.append(["P001", "JE-1", date(2026, 7, 31), 2026, 7, "개발비", "현금 지급", "대변", 1000000000, "100100", "현금"])
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        return Path(handle.name)

    def test_sheet3_mapping_normalization_event_and_risk(self) -> None:
        path = self.make_gl()
        try:
            proposal = propose_mapping(path, "GENERAL_LEDGER")
            self.assertEqual(proposal.sheet_name, "Sheet3")
            self.assertFalse(proposal.missing_required)
            company = CompanySettings("P001", "테스트회사", "소재")
            profile = MappingProfile(
                company_id=company.id,
                source_type="GENERAL_LEDGER",
                sheet_name=proposal.sheet_name,
                header_row=proposal.header_row,
                source_signature=proposal.signature,
                mapping=proposal.mapping,
                status=MappingStatus.APPROVED,
            )
            lines, reconciliation = normalize_general_ledger(path, company.id, profile)
            self.assertEqual(len(lines), 2)
            self.assertTrue(reconciliation.balanced)
            event = construct_event(lines)
            self.assertEqual(event.event_type, "DEVELOPMENT_COST_CAPITALIZATION")
            self.assertEqual(len(event.event_hash), 64)
            repo = InMemoryRepository()
            repo.save(company)
            repo.save(
                MaterialityProfile(
                    company_id=company.id,
                    name="기본",
                    benchmark="TOTAL_ASSETS",
                    overall_materiality=Decimal("500000000"),
                    performance_materiality=Decimal("300000000"),
                    trivial_threshold=Decimal("10000000"),
                    effective_from=date(2026, 1, 1),
                    status="APPROVED",
                )
            )
            result = process_journals(repo, lines, actor="test")
            self.assertEqual(result["risks"], 1)
            risk = next(iter(repo.risks.values()))
            self.assertEqual(risk.package.generated_by, "RULE_TEMPLATE")
            self.assertGreaterEqual(risk.score, 75)
        finally:
            path.unlink(missing_ok=True)


class MaterialityRepositoryTest(unittest.TestCase):
    def test_upsert_materiality_keeps_one_current_profile_per_company(self) -> None:
        repository = InMemoryRepository(persistent=False)
        company = repository.save(CompanySettings("P001", "Test company", "Manufacturing"))
        first = repository.upsert_materiality_profile(
            company.id,
            name="Default materiality",
            benchmark="REVENUE",
            overall_materiality=Decimal("500000000"),
            performance_materiality=Decimal("300000000"),
            trivial_threshold=Decimal("10000000"),
            effective_from=date(2026, 1, 1),
        )
        second = repository.upsert_materiality_profile(
            company.id,
            name="Default materiality",
            benchmark="TOTAL_ASSETS",
            overall_materiality=Decimal("600000000"),
            performance_materiality=Decimal("400000000"),
            trivial_threshold=Decimal("20000000"),
            effective_from=date(2026, 2, 1),
        )

        self.assertEqual(first.id, second.id)
        self.assertEqual(len(repository.materiality_profiles), 1)
        self.assertEqual(second.overall_materiality, Decimal("600000000"))
        self.assertEqual(repository.get_materiality_profile(company.id), second)


class VarianceTest(unittest.TestCase):
    def test_ytd_monthly_flow_and_threshold(self) -> None:
        company_id = uuid4()
        rows = [
            SettlementRow("2025-07", "410000", "매출", "REVENUE", Decimal("500"), "YTD_CUMULATIVE"),
            SettlementRow("2025-06", "410000", "매출", "REVENUE", Decimal("400"), "YTD_CUMULATIVE"),
            SettlementRow("2026-06", "410000", "매출", "REVENUE", Decimal("600"), "YTD_CUMULATIVE"),
            SettlementRow("2026-07", "410000", "매출", "REVENUE", Decimal("900"), "YTD_CUMULATIVE"),
        ]
        values = monthly_values(rows)
        self.assertEqual(values[("2026-07", "410000")], Decimal("300"))
        profile = VarianceProfile(
            company_id=company_id,
            name="기본",
            status="APPROVED",
            thresholds=[
                VarianceThreshold("MOM", Decimal("100"), Decimal("0.2"), Decimal("10"), "ANY"),
                VarianceThreshold("YOY", Decimal("100"), Decimal("0.2"), Decimal("10"), "ANY"),
            ],
        )
        observations = analyze_variance(company_id, rows, profile, "2026-07", "YOY")
        self.assertEqual(len(observations), 1)
        self.assertEqual(observations[0].current_value, Decimal("300"))
        self.assertEqual(observations[0].comparison_value, Decimal("100"))
        self.assertEqual(observations[0].measurement_basis, "MONTHLY_FLOW_FROM_YTD")


if __name__ == "__main__":
    unittest.main()
