from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.domain.models import JournalLine, MappingProfile
from app.services.mapping import UPLOAD_PERIOD_TOKEN, normalize_header


@dataclass(slots=True)
class ReconciliationResult:
    source_rows: int = 0
    accepted_rows: int = 0
    duplicate_rows: int = 0
    rejected_rows: int = 0
    debit_total: Decimal = Decimal("0")
    credit_total: Decimal = Decimal("0")
    errors: list[str] = field(default_factory=list)

    @property
    def balanced(self) -> bool:
        return self.debit_total == self.credit_total


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"invalid amount: {value!r}") from exc


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace(".", "-").replace("/", "-")
    if len(text) == 8 and text.isdigit():
        return datetime.strptime(text, "%Y%m%d").date()
    return date.fromisoformat(text[:10])


def _dc(value: Any, debit: Decimal, credit: Decimal) -> str:
    text = str(value or "").strip().upper()
    if text in {"D", "DR", "차", "차변", "DEBIT", "1"}:
        return "D"
    if text in {"C", "CR", "대", "대변", "CREDIT", "2"}:
        return "C"
    return "D" if debit and not credit else "C"


def normalize_general_ledger(
    path: str | Path,
    company_id: UUID,
    profile: MappingProfile,
    seen_hashes: set[str] | None = None,
    source_filename: str = "",
) -> tuple[list[JournalLine], ReconciliationResult]:
    if profile.status.value != "APPROVED":
        raise ValueError("approved mapping profile required")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[profile.sheet_name]
    headers = [
        str(cell.value or "").strip()
        for cell in sheet[profile.header_row]
    ]
    header_indexes = {
        normalize_header(header): index for index, header in enumerate(headers)
    }

    def value(row: tuple[Any, ...], field: str, default: Any = "") -> Any:
        source = profile.mapping.get(field)
        if not source:
            return default
        index = header_indexes.get(normalize_header(source))
        return row[index] if index is not None and index < len(row) else default

    seen = seen_hashes if seen_hashes is not None else set()
    lines: list[JournalLine] = []
    result = ReconciliationResult()
    try:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=profile.header_row + 1, values_only=True),
            start=profile.header_row + 1,
        ):
            if not any(cell not in (None, "") for cell in row):
                continue
            result.source_rows += 1
            try:
                debit = _decimal(value(row, "local_debit_amount"))
                credit = _decimal(value(row, "local_credit_amount"))
                amount = _decimal(value(row, "local_amount"))
                dc = _dc(value(row, "debit_credit_indicator"), debit, credit)
                if not amount:
                    amount = debit if dc == "D" else credit
                posting = _date(value(row, "posting_date"))
                document = str(value(row, "document_number")).strip()
                account_code = str(value(row, "account_code")).strip()
                if not document or not account_code:
                    raise ValueError("document_number and account_code are required")
                fingerprint = "|".join(
                    [
                        str(company_id),
                        source_filename,
                        document,
                        str(row_number),
                        posting.isoformat(),
                        account_code,
                        dc,
                        str(amount),
                    ]
                )
                source_hash = hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()
                if source_hash in seen:
                    result.duplicate_rows += 1
                    continue
                seen.add(source_hash)
                line = JournalLine(
                    company_id=company_id,
                    source_row=row_number,
                    document_number=document,
                    posting_date=posting,
                    account_code=account_code,
                    account_name=str(value(row, "account_name")).strip(),
                    local_amount=abs(amount),
                    debit_credit_indicator=dc,
                    fiscal_year=int(value(row, "fiscal_year", posting.year) or posting.year),
                    fiscal_period=int(value(row, "fiscal_period", posting.month) or posting.month),
                    line_text=str(value(row, "line_text")).strip(),
                    header_text=str(value(row, "header_text")).strip(),
                    project_code=str(value(row, "project_code")).strip(),
                    contract_code=str(value(row, "contract_code")).strip(),
                    vendor_code=str(value(row, "vendor_code")).strip(),
                    customer_code=str(value(row, "customer_code")).strip(),
                    source_filename=source_filename,
                    source_hash=source_hash,
                )
                lines.append(line)
                result.accepted_rows += 1
                if dc == "D":
                    result.debit_total += abs(amount)
                else:
                    result.credit_total += abs(amount)
            except Exception as exc:
                result.rejected_rows += 1
                result.errors.append(f"row {row_number}: {exc}")
    finally:
        workbook.close()
    if not result.balanced:
        result.errors.append(
            f"DEBIT_CREDIT_MISMATCH:{result.debit_total}:{result.credit_total}"
        )
    return lines, result


@dataclass(slots=True)
class SettlementRow:
    period: str
    account_code: str
    account_name: str
    category: str
    amount: Decimal
    measurement_basis: str
    current_amount: Decimal = Decimal("0")
    prior_amount: Decimal = Decimal("0")

    def __post_init__(self) -> None:
        # Historical variance rows were persisted with a single amount column.
        if self.current_amount == Decimal("0") and self.amount != Decimal("0"):
            self.current_amount = self.amount


def normalize_settlement(
    path: str | Path,
    profile: MappingProfile,
    *,
    upload_period: str | None = None,
) -> list[SettlementRow]:
    if profile.status.value != "APPROVED":
        raise ValueError("approved mapping profile required")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[profile.sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[profile.header_row]]
    indexes = {normalize_header(header): index for index, header in enumerate(headers)}

    def value(row: tuple[Any, ...], field: str, default: Any = "") -> Any:
        source = profile.mapping.get(field)
        if source == UPLOAD_PERIOD_TOKEN:
            if not upload_period:
                raise ValueError("upload_period is required when the settlement file has no period column")
            return upload_period
        index = indexes.get(normalize_header(source)) if source else None
        return row[index] if index is not None else default

    rows: list[SettlementRow] = []
    try:
        for row in sheet.iter_rows(min_row=profile.header_row + 1, values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            category = str(value(row, "category", "OTHER")).strip().upper()
            basis = str(value(row, "measurement_basis")).strip().upper()
            if not basis:
                basis = "YTD_CUMULATIVE" if category in {"REVENUE", "EXPENSE", "수익", "비용"} else "ENDING_BALANCE"
            rows.append(
                SettlementRow(
                    period=str(value(row, "period")).strip(),
                    account_code=str(value(row, "account_code")).strip(),
                    account_name=str(value(row, "account_name")).strip(),
                    category=category,
                    amount=_decimal(value(row, "current_amount")) if profile.mapping.get("current_amount") else _decimal(value(row, "amount")),
                    measurement_basis=basis,
                    current_amount=_decimal(value(row, "current_amount")) if profile.mapping.get("current_amount") else _decimal(value(row, "amount")),
                    prior_amount=_decimal(value(row, "prior_amount")),
                )
            )
    finally:
        workbook.close()
    return rows
