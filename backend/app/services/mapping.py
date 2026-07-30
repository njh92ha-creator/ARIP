from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.domain.models import MappingProfile


# A settlement schedule commonly has a single current-period column (for
# example, "당기") instead of separate period and amount columns.  The period
# is supplied by the user at AVI execution time in that case.
UPLOAD_PERIOD_TOKEN = "__UPLOAD_PERIOD__"


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "company_code": ("회사코드", "법인코드", "companycode"),
    "document_number": ("전표번호", "문서번호", "documentno", "journalno"),
    "posting_date": ("전기일자", "전표일자", "회계일자", "postingdate"),
    "fiscal_year": ("회계연도", "연도", "fiscalyear"),
    "fiscal_period": ("회계기간", "기간", "월", "fiscalperiod"),
    "header_text": ("전표적요", "헤더적요", "documenttext"),
    "line_text": ("전표적요상세", "적요", "비고", "linetext", "description"),
    "debit_credit_indicator": ("차대구분", "차대", "차변대변구분", "debitcredit"),
    "local_amount": ("전표금액기준통화", "전표금액", "금액", "localamount"),
    "local_debit_amount": ("차변금액기준통화", "차변금액", "debitamount"),
    "local_credit_amount": ("대변금액기준통화", "대변금액", "creditamount"),
    "account_code": ("계정과목코드", "계정코드", "accountcode"),
    "account_name": ("계정과목명", "계정명", "accountname"),
    "project_code": ("프로젝트코드", "프로젝트", "projectcode"),
    "contract_code": ("계약번호", "계약코드", "contractcode"),
    "vendor_code": ("구매처코드", "공급업체코드", "vendorcode"),
    "customer_code": ("고객코드", "customer_code", "customercode"),
    "period": ("기준월", "회계기간", "기준기간", "월", "period"),
    "category": ("재무제표구분", "계정분류", "구분", "category"),
    "amount": (
        "금액", "당기", "당월", "당기말", "기말잔액", "누계", "누적금액",
        "잔액", "amount",
    ),
    "measurement_basis": ("측정기준", "금액기준", "basis"),
}

REQUIRED_FIELDS = {
    "GENERAL_LEDGER": {"document_number", "posting_date", "account_code"},
    # A settlement schedule can use UPLOAD_PERIOD_TOKEN in place of a period
    # column.  The AVI execution form supplies that period explicitly.
    "SETTLEMENT_SCHEDULE": {"account_code", "amount"},
}

SOURCE_FIELDS = {
    "GENERAL_LEDGER": {
        "company_code", "document_number", "posting_date", "fiscal_year",
        "fiscal_period", "header_text", "line_text", "debit_credit_indicator",
        "local_amount", "local_debit_amount", "local_credit_amount", "account_code",
        "account_name", "project_code", "contract_code", "vendor_code", "customer_code",
    },
    "SETTLEMENT_SCHEDULE": {
        "period", "account_code", "account_name", "category", "amount", "measurement_basis",
    },
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s\(\)\[\]_\-/,.]", "", str(value or "")).lower()


def source_signature(headers: Iterable[Any]) -> str:
    payload = json.dumps(
        sorted(normalize_header(header) for header in headers if header is not None),
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


@dataclass(slots=True)
class MappingProposal:
    sheet_name: str
    header_row: int
    headers: list[str]
    mapping: dict[str, str]
    confidence: dict[str, float]
    missing_required: list[str]
    signature: str


def _score(source: str, aliases: tuple[str, ...]) -> float:
    normalized = normalize_header(source)
    if not normalized:
        return 0.0
    best = 0.0
    for alias in aliases:
        candidate = normalize_header(alias)
        if normalized == candidate:
            return 1.0
        if candidate and (candidate in normalized or normalized in candidate):
            best = max(best, 0.82)
    return best


def _detect_header(sheet: Any, max_rows: int = 25) -> tuple[int, list[str]]:
    best_row, best_headers, best_hits = 1, [], -1
    aliases = tuple(alias for values in FIELD_ALIASES.values() for alias in values)
    for row_index in range(1, min(max_rows, sheet.max_row) + 1):
        headers = [str(cell.value or "").strip() for cell in sheet[row_index]]
        hits = sum(
            1 for header in headers if max((_score(header, (a,)) for a in aliases), default=0) >= 0.8
        )
        if hits > best_hits:
            best_row, best_headers, best_hits = row_index, headers, hits
    return best_row, best_headers


def propose_mapping(
    path: str | Path,
    source_type: str,
    *,
    force_sheet: str | None = None,
) -> MappingProposal:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if source_type == "GENERAL_LEDGER":
            sheet_name = force_sheet or ("Sheet3" if "Sheet3" in workbook.sheetnames else workbook.sheetnames[0])
        else:
            sheet_name = force_sheet or workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        header_row, headers = _detect_header(sheet)
        mapping: dict[str, str] = {}
        confidence: dict[str, float] = {}
        for field_name in SOURCE_FIELDS[source_type]:
            aliases = FIELD_ALIASES[field_name]
            candidates = [(header, _score(header, aliases)) for header in headers if header]
            if candidates:
                header, score = max(candidates, key=lambda item: item[1])
                if score >= 0.70:
                    mapping[field_name] = header
                    confidence[field_name] = score

        if source_type == "SETTLEMENT_SCHEDULE" and "period" not in mapping:
            mapping["period"] = UPLOAD_PERIOD_TOKEN
            confidence["period"] = 1.0

        required = REQUIRED_FIELDS[source_type]
        if source_type == "GENERAL_LEDGER":
            amount_present = "local_amount" in mapping or {
                "local_debit_amount", "local_credit_amount"
            }.issubset(mapping)
            missing = sorted(required - set(mapping))
            if not amount_present:
                missing.append("local_amount OR debit/credit amounts")
        else:
            missing = sorted(required - set(mapping))
        return MappingProposal(
            sheet_name=sheet_name,
            header_row=header_row,
            headers=headers,
            mapping=mapping,
            confidence=confidence,
            missing_required=missing,
            signature=source_signature(headers),
        )
    finally:
        workbook.close()


def validate_mapping(profile: MappingProfile, headers: list[str]) -> list[str]:
    errors: list[str] = []
    normalized_headers = {normalize_header(header) for header in headers}
    if source_signature(headers) != profile.source_signature:
        errors.append("SCHEMA_DRIFT")
    for field_name, source_column in profile.mapping.items():
        if source_column == UPLOAD_PERIOD_TOKEN:
            continue
        if normalize_header(source_column) not in normalized_headers:
            errors.append(f"MISSING_COLUMN:{field_name}:{source_column}")
    return errors
